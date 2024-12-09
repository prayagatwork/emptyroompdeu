import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from pymongo import MongoClient
import os
import xlrd

filename = None

def browse_file():
    global filename
    f_name = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx")])
    if f_name:
        filename = f_name
        entry.delete(0, tk.END)  # Clear the entry box
        entry.insert(tk.END, filename)

def process_file():
    global filename
    if filename:
        # Connect to MongoDB
        client = MongoClient('mongodb+srv://prayagadmin:me1234@pdeuer.p8bjgio.mongodb.net/?retryWrites=true&w=majority&appName=pdeuER')
        db = client['ListOfEmpty']
        collection = db['empty-rooms']
        
        block_name = os.path.split(filename)[-1][0]

        def find_rnums(sheet):
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).startswith(block_name + "-"):
                        yield cell.value

        rooms_list_f = []
        workbook = load_workbook(filename)
        sheet = workbook.active

        # Extract room numbers
        for room in find_rnums(sheet):
            rooms_list_f.append(room)

        
        # Remove block prefix and store room numbers
        rooms_list = [x.lstrip(block_name ) for x in rooms_list_f]

        total_rows = sheet.max_row
        total_columns = sheet.max_column
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        rooms_schedule = {}  # Dictionary to store room schedules

        row_no = 0
        rooms_marker = 0

        row = list(sheet.iter_rows(max_row = total_rows, max_col = total_columns))
        while row_no < total_rows:
            first_cell_value = row[row_no][0].value
            if first_cell_value and first_cell_value.lower() == "monday":
                room_schedule = {}  # Dictionary to store schedule for a room
                for d in range(5):
                    i = 2
                    if d < 4:
                        next_day = days[d + 1]
                    temp = []

                    while i<=10:
                        merged_horizontal = False

                        # Check if the cell is part of a merged range
                        if row[row_no][i].coordinate in sheet.merged_cells:                   
                            # Get the merged range
                            for merged_range in sheet.merged_cells.ranges:
                                if row[row_no][i].coordinate in merged_range:
                                    # Check if the merged range spans horizontally
                                    if merged_range.max_col > merged_range.min_col:
                                        # Set the flag indicating horizontal merge
                                        merged_horizontal = True
                        if merged_horizontal and row[row_no][i].value is not None: #for filled lab
                            i+=2 #considering every lab is of 2 hours

                        elif merged_horizontal and row[row_no][i].value is None: #for empty lab
                            if i>=6:
                                temp.append(str(i-5))
                                temp.extend(str(i+1-5))
                            else:
                                temp.extend(str(i+7))
                                temp.extend(str(i+7+1))
                            i+=2

                        elif row[row_no][i].value is not None:
                            i+=1

                        elif row[row_no][i].value is None:
                            if i>=6:
                                temp.append(str(i-5))
                            else:
                                temp.append(str(i+7))
                            i+=1

                    if d < 4:
                        while row[row_no][0].value != next_day:
                            row_no += 1

                    room_schedule[days[d]] = temp

                room_number = rooms_list[rooms_marker]
                rooms_schedule[room_number] = room_schedule
                if rooms_marker<len(rooms_list)-1:
                    rooms_marker+=1
                elif rooms_marker == len(rooms_list)-1:
                    row_no = total_rows
            else:
                row_no += 1

        # Construct document

    initial_document = {
            "block": str(block_name),
            "rooms": rooms_schedule
        }


    rooms_list_f = []
    for room_number, schedule in initial_document['rooms'].items():
        room_dict_f = {"room_number": room_number, "schedule": schedule}
        rooms_list_f.append(room_dict_f)


    final_document = {
        "block": initial_document["block"],
        "rooms": rooms_list_f
    }


        

    print(final_document)  
    
    collection.insert_one(final_document)
    print("Document inserted into MongoDB collection.")
    client.close()



def ok_action():
    process_file()
    root.destroy()

root = tk.Tk()
root.title("Excel File Processor")

label = tk.Label(root, text="Enter File Path:")
label.pack()

entry = tk.Entry(root, width=50)
entry.pack()

browse_button = tk.Button(root, text="Browse", command=browse_file)
browse_button.pack()

ok_button = tk.Button(root, text="OK", command=ok_action)
ok_button.pack()

root.mainloop()